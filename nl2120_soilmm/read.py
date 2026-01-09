import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import numpy as np

from nl2120_soilmm.constants import (
    GROUNDWATER_WELLS,
    EXTENSOMETER_DEPTHS,
    SOILTYPES_ENGLISH,
    SOILTYPES_DUTCH_REGIODEAL,
    HYDRAULIC_HEADS,
    MIDDEPTH_FILTERS,
    LOCATION_FULLNAMES,
    DITCHES,
    SELECTED_GROUNDWATER_WELLS,
)


def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames


def read_hydraulic_head(location, plot_type="RF"):
    """
    Reads hydraulic head data for a given location and plot type.
    Functionality to select the plot type still needs to be implemented.
    """

    hydraulic_head = HYDRAULIC_HEADS[location]
    location_fullname = LOCATION_FULLNAMES[location]

    match location:
        case "LW":
            foldername = "LAW"
        case "ZH":
            foldername = "ZEG31"
        case _:
            foldername = location

    match location:
        case "GDA" | "BKW" | "BKG" | "CBW" | "HZW" | "MMW" | "M4T" | "MSW":
            basedir = Path(
                rf"n:/Projects/11211000/11211391/B. Measurements and calculations/Bodembeweging/data/2-interim/{location_fullname}"
            )
        case _:
            # groundwater data
            basedir = Path(
                f"p:/broeikasgassen-veenweiden/Grondwater/grondwaterstandanalyse/data/4-output/Gecorrigeerde_grondwaterstanden_hourly/gecorrigeerd/{foldername}"
            )

    if hydraulic_head:
        data = pd.read_csv(
            basedir.joinpath(f"{hydraulic_head}.csv"), index_col=0, parse_dates=True
        )["Waterstand"]

        return data
    else:
        pass


def read_ditch_level(location, plot_type="RF"):
    """
    Reads hydraulic head data for a given location and plot type.
    Functionality to select the plot type still needs to be implemented.
    """

    match location:
        case "GDA" | "BKW" | "BKG" | "CBW" | "HZW" | "MMW" | "M4T" | "MSW":
            ditch = DITCHES[location]
        case _:
            ditch = DITCHES[f"{location} {plot_type}"]
    location_fullname = LOCATION_FULLNAMES[location]

    match location:
        case "LW":
            foldername = "LAW"
        case "ZH":
            foldername = "ZEG31"
        case _:
            foldername = location

    match location:
        case "GDA" | "BKW" | "BKG" | "CBW" | "HZW" | "MMW" | "M4T" | "MSW":
            basedir = Path(
                rf"n:/Projects/11211000/11211391/B. Measurements and calculations/Bodembeweging/data/2-interim/{location_fullname}"
            )
        case _:
            # groundwater data
            basedir = Path(
                f"p:/broeikasgassen-veenweiden/Grondwater/grondwaterstandanalyse/data/4-output/Gecorrigeerde_grondwaterstanden_hourly/gecorrigeerd/{foldername}"
            )

    if ditch:
        data = pd.read_csv(
            basedir.joinpath(f"{ditch}.csv"), index_col=0, parse_dates=True
        )["Waterstand"]

        return data
    else:
        pass


def read_precipitation_deficit(location):

    basedir = Path(
        "n:/Projects/11211000/11211391/B. Measurements and calculations/Bodembeweging/data"
    )

    path_to_data = basedir.joinpath(
        "2-interim",
        LOCATION_FULLNAMES[location],
        f"{location}_precipitation_deficit.csv",
    )

    data = pd.read_csv(path_to_data, index_col=0, parse_dates=True)

    # convert the dataframe to a series by selecting the first column
    data = data.iloc[:, 0]

    return data


def read_extensometer(location, plot_type="RF"):

    match location:
        case (
            "DEM"
            | "LW"
            | "VEG"
            | "ZH"
            | "GDA"
            | "BKW"
            | "BKG"
            | "CBW"
            | "HZW"
            | "MMW"
            | "M4T"
            | "MSW"
            | "HGM"
            | "HGG"
            | "HGR"
        ):
            extensometer_depth = EXTENSOMETER_DEPTHS[f"{location}"]
            filename_plot_type = f"{location}_extensometer.csv"
        case _:
            extensometer_depth = EXTENSOMETER_DEPTHS[f"{location} {plot_type}"]
            filename_plot_type = f"{location}_extensometer_{plot_type}.csv"

    basedir = Path(
        "n:/Projects/11211000/11211391/B. Measurements and calculations/Bodembeweging/data"
    )

    path_to_data = basedir.joinpath(
        "2-interim", LOCATION_FULLNAMES[location], filename_plot_type
    )

    data = pd.read_csv(path_to_data, index_col=0, parse_dates=True)[extensometer_depth]
    data.index = pd.to_datetime(data.index)

    column_names = [column_name.replace("m-mv", "m bs") for column_name in data.columns]

    # hier gebleven, kolom naam kan nog niet aangepast worden
    match location:
        case "ROU" | "ROU09":
            column_names[-1] = column_names[-1].replace("referentie ", "")

    data.columns = column_names

    # multiply the anchor level in the column name by 100 to get the anchor level in cm
    for column in data.columns:
        anchor_level = float(column[:4])
        data.rename(columns={column: f"{anchor_level * 100:0.0f} cm bs"}, inplace=True)

    if location == "HZW":
        data = data.drop(columns=["260 cm bs"])

    if location == "ROU":
        if plot_type == "MS":
            data.loc["2025-08-19 14:00":, "115 cm bs"] = np.nan

    # divide by 10 to get the data in cm
    data_cm = data / 10

    # the frequency of the data is hourly, so explicitly set it to hourly
    data_cm = data_cm.asfreq("h")

    # data_cm.interpolate(method="linear", inplace=True)

    return data_cm


def read_surface_level(location, plot_type="RF"):

    basedir = Path(
        "n:/Projects/11211000/11211391/B. Measurements and calculations/Bodembeweging/data"
    )

    match location:
        case (
            "DEM"
            | "LW"
            | "VEG"
            | "ZH"
            | "GDA"
            | "BKW"
            | "BKG"
            | "CBW"
            | "HZW"
            | "MMW"
            | "M4T"
            | "MSW"
            | "HGM"
            | "HGG"
            | "HGR"
        ):
            plot_type = ""
        case _:
            plot_type = "_" + plot_type

    path_to_data = basedir.joinpath(
        "2-interim",
        LOCATION_FULLNAMES[location],
        f"{location}_surface_level{plot_type}.csv",
    )

    data = pd.read_csv(path_to_data, header=None)

    # multiply by 100 to get the data in cm
    data_cm = data * 100

    return data_cm.iloc[0, 0]


def read_filter_depths(location, plot_type="RF"):

    basedir = Path(
        "n:/Projects/11211000/11211391/B. Measurements and calculations/Bodembeweging/data"
    )

    match location:
        case (
            "DEM"
            | "LW"
            | "VEG"
            | "ZH"
            | "GDA"
            | "BKW"
            | "BKG"
            | "CBW"
            | "HZW"
            | "MMW"
            | "M4T"
            | "MSW"
        ):
            plot_type = ""
        case _:
            plot_type = "_" + plot_type

    path_to_data = basedir.joinpath(
        "2-interim",
        LOCATION_FULLNAMES[location],
        f"{location}_filterdepths{plot_type}.csv",
    )

    data = pd.read_csv(path_to_data, header=None)

    # multiply by 100 to get the data in cm
    data_cm = data * 100

    return data_cm


def read_soilprofile(location, location_fullname, plot_type="RF", language="english"):

    match location:
        case "ROU":
            if plot_type == "MS":
                sheetname = "ROU05_MP"
            else:
                sheetname = "ROU05_RF"
        case "ALB" | "ASD" | "VLI" | "ZEG":
            sheetname = f"{location}_{plot_type}"
        case "DEM" | "VEG":
            sheetname = location
        case "LW":
            sheetname = "LAW"
        case "ZH":
            sheetname = "ZEG_HW"
        case "BKW" | "BKG" | "CBW":
            sheetname = location_fullname
        case "HZW":
            sheetname = location_fullname.removesuffix("-Dorp")
        case "ROU09":
            if plot_type == "MS":
                sheetname = "ROU09_MP"
            else:
                sheetname = "ROU09_RF"
        case "M4T":
            sheetname = "Vierde tochtweg"
        case "MMW":
            sheetname = "Middelweg"
        case "MSW":
            sheetname = "Spoorweglaan"
        case "HGM":
            sheetname = "Museumsite"
        case "HGG":
            sheetname = "Perceel 3 - greppel"
        case "HGR":
            sheetname = "Perceel 3 - referentie"

    if location == "ZEG" and plot_type == "MS":
        sheetname = "ZEG_003"

    basedir = Path(
        r"N:/Projects/11211000/11211391/B. Measurements and calculations/Bodembeweging"
    )

    match location:
        case "GDA" | "BKW" | "BKG" | "CBW" | "HZW":
            filepath = basedir.joinpath(
                "data/3-processed/regiodeal/lithologie en ankerdiepten RDBGH44.xlsx"
            )
        case "M4T" | "MMW" | "MSW":
            filepath = basedir.joinpath(
                "data/3-processed/Lithologie en ankerdiepten restveengebied.xlsx"
            )
        case "HGM" | "HGG" | "HGR":
            filepath = basedir.joinpath(
                "data/3-processed/Lithologie en ankerdiepten Hegewarren.xlsx"
            )
        case _:
            filepath = basedir.joinpath(
                "data/3-processed/Lithologie en ankerdiepten extensometers aangevuld.xlsx"
            )

    soil_profile = pd.read_excel(
        filepath,
        sheet_name=sheetname,
    )

    lithology = (
        soil_profile[["lithologie", "bovengrens [cm]", "ondergrens [cm]"]]
        .dropna()
        .set_index("lithologie")
    )

    lithology = lithology.astype(
        {"bovengrens [cm]": "float64", "ondergrens [cm]": "float64"}
    )

    lithology.loc[:, ("bovengrens [cm]", "ondergrens [cm]")] = lithology.loc[
        :, ("bovengrens [cm]", "ondergrens [cm]")
    ]
    lithology["dikte"] = lithology.iloc[:, 1] - lithology.iloc[:, 0]

    if language != "dutch":
        lithology.index = [SOILTYPES_ENGLISH.get(l[0]) for l in lithology.iterrows()]

    anchors = (
        soil_profile[["anker", "m-mv", "m NAP"]].iloc[::-1].set_index("anker").dropna()
    )

    # anchors['m-mv'] *= -1 # to obtain positive values for m-mv

    # reverse the order of the soil profile anchors to match the order of the extensometer data
    match location:
        case "DEM" | "LW" | "VEG":
            anchors = anchors.iloc[::-1]

    # Remove the anchor with m-mv = -2.60 for Hazerswoude, as it is not used in the analysis
    if location == "HZW":
        anchors = anchors[anchors["m-mv"] != -2.60]

    return lithology, anchors


def read_soilprofile_regiodeal(location, location_fullname):

    try:
        path_to_soilprofile = Path(
            f"n:/Projects/11211000/11211391/B. Measurements and calculations/Bodembeweging/data/3-processed/regiodeal/{location}_lithology.csv"
        )
        soilprofile = pd.read_csv(
            path_to_soilprofile,
            index_col=0,
        )

        lithology = (
            soilprofile[["lithology", "humusAdmix"]].fillna("").agg("".join, axis=1)
        )
        lithology.name = "lithology"

        lithology = lithology.to_frame()

        lithology["bovengrens [cm]"] = soilprofile["top"]
        lithology["ondergrens [cm]"] = soilprofile["bottom"]

        lithology = lithology.set_index("lithology")
        lithology.index = [
            SOILTYPES_DUTCH_REGIODEAL.get(l[0]) for l in lithology.iterrows()
        ]

        lithology = lithology.astype(
            {"bovengrens [cm]": "float64", "ondergrens [cm]": "float64"}
        )

        lithology["dikte"] = lithology["ondergrens [cm]"] - lithology["bovengrens [cm]"]
    except FileNotFoundError:
        print(f"Soilprofile not found for location {location_fullname}.")
        lithology = pd.DataFrame()

    # read the anchors
    path_to_anchors = Path(
        f"n:/Projects/11211000/11211391/B. Measurements and calculations/Bodembeweging/data/3-processed/regiodeal/{location}_anchors.csv"
    )
    anchors = pd.read_csv(
        path_to_anchors, index_col=0, usecols=["anker", "m-mv", "m NAP"]
    ).iloc[::-1]

    return lithology, anchors


def read_strain(path_stats, sheetname, header=2, footer=0):

    strain_values = pd.read_excel(
        path_stats,
        sheet_name=sheetname,
        header=header,
        usecols="F, I, K",
        index_col=0,
        skipfooter=footer,
    )

    strain_values.dropna(inplace=True)

    lower_bounds = []
    upper_bounds = []
    for depth in strain_values.index:
        lower_bounds.append(float(depth.split(" ")[0]))
        upper_bounds.append(float(depth.split(" ")[-3]))

    strain_values["lower bounds"] = lower_bounds
    strain_values["upper bounds"] = upper_bounds

    return strain_values


def read_gwlevel(location, plot_type):

    location_fullname = LOCATION_FULLNAMES[location]

    match location:
        case "LW":
            foldername = "LAW"
        case "ZH":
            foldername = "ZEG31"
        case _:
            foldername = location

    match location:
        case "GDA" | "BKW" | "BKG" | "CBW" | "HZW" | "ROU09" | "MMW" | "M4T" | "MSW":
            basedir = Path(
                rf"n:/Projects/11211000/11211391/B. Measurements and calculations/Bodembeweging/data/2-interim/{location_fullname}"
            )

            data = pd.read_csv(
                basedir.joinpath(f"{location}_gwlevels.csv"),
                index_col=0,
                parse_dates=True,
            )  # ["Waterstand"]

        case _:
            # groundwater data
            basedir = Path(
                f"p:/broeikasgassen-veenweiden/Grondwater/grondwaterstandanalyse/data/4-output/Gecorrigeerde_grondwaterstanden_hourly/gecorrigeerd/{foldername}"
            )

            data = pd.DataFrame()

            wells = SELECTED_GROUNDWATER_WELLS[f"{location} {plot_type}"]

            for well in wells:
                if well:
                    data_single_well = pd.read_csv(
                        basedir.joinpath(f"{well}.csv"), index_col=0, parse_dates=True
                    )["Waterstand"]

                else:
                    data_single_well = pd.DataFrame()

                data = pd.concat([data, data_single_well], axis=1)

            data.columns = [well for well in wells if well]

    return data


if __name__ == "__main__":

    from nl2120_soilmm.constants import LOCATION_FULLNAMES

    location = "MMW"

    location_fullname = LOCATION_FULLNAMES[location]

    # locations = ["ALB", "ASD", "ROU", "VLI", "VEG", "DEM", "LW", "ZH", "ZEG"]

    # for location in locations:
    print(f"Processing data for location: {location_fullname}")

    #####################################
    # test inidividual functions
    #####################################
    # groundwater_data = read_groundwater(location)
    # hydrolic_head_data = read_hydraulic_head(location)
    # extensometer_data = read_extensometer(location)
    # middepth_filter_data = read_middepth_filter(location)
    # surface_level = read_surface_level(location)
    # precip_deficit = read_precipitation_deficit(location)
    # filter_depths = read_filter_depths_phreatic(location)

    lithology, anchor = read_soilprofile(location, location_fullname)
    # lithology, anchors = read_soilprofile_regiodeal(location, location_fullname)

    # if location in ["ZEG", "ROU", "ALB", "VLI", "ASD"]:
    #     path_stats = r"n:/Projects/11211000/11211391/B. Measurements and calculations/Bodembeweging/data/4-output/trendlijn_statistieken.xlsx"
    # elif location in ["GDA", "HZW", "BKG", "BKW", "CBW"]:
    #     path_stats = r"n:/Projects/11211000/11211391/B. Measurements and calculations/Bodembeweging/data/4-output/trendlijn_statistieken_regiodeal.xlsx"

    # strain = read_strain(path_stats, sheetname=location_fullname, location=location)
