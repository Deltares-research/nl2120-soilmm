from pathlib import Path
from openpyxl import load_workbook
import pandas as pd


def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames


def write_precipitation_deficit(location, location_fullname):
    basedir = Path(
        "n:/Projects/11204000/11204108/B. Measurements and calculations/Meetlocaties/Info_per_locatie/Zegveld/zeg_pt"
    )

    path_to_data = basedir.joinpath("zeg_meteo_data.xlsx")

    sheetnames = get_sheetnames_xlsx(path_to_data)
    print(sheetnames)
    years = sheetnames[1:-1]

    data = pd.DataFrame()

    for year in years:
        data_year = pd.read_excel(
            path_to_data,
            sheet_name=year,
            index_col=0,
            parse_dates=True,
            usecols="A,AJ",
            # usecols=["TIMESTAMP_END", "P_deficit"],
        )

        if data.empty:
            data = data_year
        else:
            data = pd.concat([data, data_year])

    data = data.resample("h").mean()

    # convert the dataframe to a series by selecting the first column
    data = data.iloc[:, 0]

    outputdir = Path(
        r"n:/Projects/11211000/11211391/B. Measurements and calculations/Bodembeweging/data/2-interim"
    )

    data.to_csv(
        outputdir.joinpath(location_fullname, f"{location}_precipitation_deficit.csv"),
    )


if __name__ == "__main__":

    LOCATION_FULLNAMES = {
        "ALB": "Aldeboarn",
        "ASD": "Assendelft",
        "ROU": "Rouveen",
        "VLI": "Vlist",
        "ZEG": "Zegveld",
        "DEM": "Demmerik",
        "LW": "LangeWeide",
        "VEG": "Vegelinsoord",
        "ZH": "ZegveldHoogwater",
        "LR": "LangRoggebroek",
    }
    location = "ZEG"
    location_fullname = LOCATION_FULLNAMES[location]

    write_precipitation_deficit(location, location_fullname)
